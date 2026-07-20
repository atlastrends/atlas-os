# ============================================================
# ATLAS OS - amazon_report_service.py
# Le os relatorios de afiliado que o usuario baixa no Amazon
# Associates (CSV ou Excel) e transforma em linhas normalizadas
# para a tabela amazon_sales. Tambem calcula as estatisticas da
# pagina "Vendas Amazon".
#
# A Amazon NAO oferece API de ganhos/vendas: por isso o usuario
# baixa o relatorio no site e importa aqui.
# ============================================================

from __future__ import annotations

import csv
import hashlib
import io
import os
import re
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.amazon_sales import AmazonSale

# Tags de afiliado por mercado (para deduzir BR/US pelo tracking id).
_TAG_TO_MARKET = {
    "achadosatlasb": "BR",
    "atlasfindsus": "US",
}
_CURRENCY_BY_MARKET = {"BR": "BRL", "US": "USD"}


# ----------------------------------------------------------------
# Normalizacao de texto de cabecalho
# ----------------------------------------------------------------
def _norm(text: Any) -> str:
    s = str(text or "").strip().lower()
    s = s.replace("ç", "c").replace("ã", "a").replace("á", "a")
    s = s.replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return s.strip()


# Palavras-chave (ja normalizadas) que identificam cada coluna.
_FIELD_KEYWORDS = {
    "product_name": ["name", "title", "produto", "nome", "product"],
    "asin": ["asin"],
    "category": ["category", "categoria"],
    "qty": [
        "items shipped", "qty shipped", "quantity", "items ordered",
        "unidades", "quantidade", "qtd", "itens enviados", "itens",
    ],
    "returns": ["returns", "devolucoes", "devolucao", "returned"],
    "revenue": [
        "revenue", "product sales", "receita", "vendas de produtos",
        "valor", "sale amount", "ordered revenue", "price",
    ],
    "commission": [
        "ad fees", "fees", "earnings", "comissao", "comissoes",
        "taxa de publicidade", "receita de publicidade", "ganhos",
        "advertising fees", "commission",
    ],
    "clicks": ["clicks", "cliques"],
    "date": ["date shipped", "date", "data", "period", "periodo", "day"],
    "tracking_id": ["tracking id", "tracking", "id de rastreamento", "rastreamento"],
}

# Ordem de prioridade para casar cabecalhos (mais especifico primeiro).
_FIELD_ORDER = [
    "asin", "tracking_id", "clicks", "commission", "returns",
    "qty", "revenue", "category", "product_name", "date",
]


def _match_headers(headers: list[str]) -> dict[str, int]:
    """Descobre em qual coluna esta cada campo, pelo nome do cabecalho."""
    normed = [_norm(h) for h in headers]
    used: set[int] = set()
    mapping: dict[str, int] = {}
    for field in _FIELD_ORDER:
        keywords = _FIELD_KEYWORDS[field]
        best_idx = -1
        best_len = 0
        for idx, h in enumerate(normed):
            if idx in used or not h:
                continue
            for kw in keywords:
                if kw in h and len(kw) > best_len:
                    best_idx = idx
                    best_len = len(kw)
        if best_idx >= 0:
            mapping[field] = best_idx
            used.add(best_idx)
    return mapping


def _looks_like_header(row: list[str]) -> bool:
    """Uma linha e cabecalho se casar pelo menos 2 campos conhecidos."""
    m = _match_headers(row)
    return len(m) >= 2


# ----------------------------------------------------------------
# Conversao de numeros (aceita formato BR e US)
# ----------------------------------------------------------------
def _parse_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return 0.0
    neg = "(" in s and ")" in s  # contabil: (1,23) = negativo
    # Mantem apenas digitos, virgula, ponto e sinal.
    s = re.sub(r"[^0-9,.\-]", "", s)
    if not s or s in {"-", ".", ","}:
        return 0.0
    if "," in s and "." in s:
        # O separador que aparece por ultimo e o decimal.
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")  # BR: 1.234,56
        else:
            s = s.replace(",", "")                     # US: 1,234.56
    elif "," in s:
        # So virgula: decimal se ate 2 digitos depois; senao milhar.
        after = s.split(",")[-1]
        if len(after) in (1, 2):
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    try:
        num = float(s)
    except ValueError:
        return 0.0
    return -num if neg else num


def _parse_int(value: Any) -> int:
    return int(round(_parse_number(value)))


def _parse_date(value: Any) -> str:
    """Devolve AAAA-MM-DD (ou string vazia se nao reconhecer)."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return ""
    # Formatos mais comuns nos relatorios da Amazon.
    fmts = [
        "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%d/%m/%y",
        "%B %d, %Y", "%d %b %Y", "%Y/%m/%d", "%m-%d-%Y",
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Tenta achar AAAA-MM-DD dentro do texto.
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    return ""


def _market_from_tracking(tracking_id: str, fallback: str) -> str:
    t = (tracking_id or "").lower()
    for tag, mkt in _TAG_TO_MARKET.items():
        if tag in t:
            return mkt
    return fallback


# ----------------------------------------------------------------
# Leitura bruta do arquivo -> lista de linhas (list[list[str]])
# ----------------------------------------------------------------
def _read_rows(filename: str, data: bytes) -> list[list[str]]:
    name = (filename or "").lower()
    is_xlsx = name.endswith(".xlsx") or data[:2] == b"PK"
    if is_xlsx:
        return _read_xlsx(data)
    return _read_csv(data)


def _read_xlsx(data: bytes) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise ValueError(
            "Nao consigo ler Excel (openpyxl ausente). "
            "Baixe o relatorio como CSV ou instale o openpyxl."
        ) from exc
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else c for c in row])
    wb.close()
    return rows


def _read_csv(data: bytes) -> list[list[str]]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = data.decode("utf-8", errors="replace")
    # Descobre o separador (virgula, ponto-e-virgula ou tab).
    sample = "\n".join(text.splitlines()[:20])
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        counts = {d: sample.count(d) for d in [",", ";", "\t"]}
        delimiter = max(counts, key=counts.get) or ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [list(r) for r in reader]


# ----------------------------------------------------------------
# API principal: parse do arquivo -> linhas normalizadas (dicts)
# ----------------------------------------------------------------
def parse_report(
    filename: str,
    data: bytes,
    default_market: str = "BR",
) -> list[dict]:
    rows = _read_rows(filename, data)
    if not rows:
        return []

    # Acha a linha de cabecalho (relatorios tem texto antes da tabela).
    header_idx = -1
    for i, row in enumerate(rows[:30]):
        if _looks_like_header([str(c) for c in row]):
            header_idx = i
            break
    if header_idx < 0:
        raise ValueError(
            "Nao reconheci as colunas do relatorio. Confirme que e um "
            "relatorio de afiliado da Amazon (Ganhos, Pedidos ou Cliques)."
        )

    headers = [str(c) for c in rows[header_idx]]
    mapping = _match_headers(headers)
    default_market = (default_market or "BR").upper()
    if default_market not in ("BR", "US"):
        default_market = "BR"

    out: list[dict] = []
    for raw in rows[header_idx + 1:]:
        if not any(str(c).strip() for c in raw):
            continue  # linha vazia

        def cell(field: str) -> Any:
            idx = mapping.get(field, -1)
            return raw[idx] if 0 <= idx < len(raw) else ""

        product_name = str(cell("product_name") or "").strip()[:500]
        asin = str(cell("asin") or "").strip()[:20]
        tracking_id = str(cell("tracking_id") or "").strip()[:80]
        category = str(cell("category") or "").strip()[:200]
        qty = _parse_int(cell("qty"))
        returns = _parse_int(cell("returns"))
        revenue = _parse_number(cell("revenue"))
        commission = _parse_number(cell("commission"))
        clicks = _parse_int(cell("clicks"))
        sale_date = _parse_date(cell("date"))

        # Linha sem produto/ASIN e quase sempre um somatorio ("Total",
        # subtotal, rodape). Uma linha real de produto sempre tem nome ou ASIN.
        if not asin and not product_name:
            continue
        low_name = _norm(product_name)
        if low_name in {"total", "totals", "grand total", "totais"}:
            continue

        market = _market_from_tracking(tracking_id, default_market)
        currency = _CURRENCY_BY_MARKET.get(market)

        if commission:
            report_type = "earnings"
        elif clicks and not revenue:
            report_type = "clicks"
        else:
            report_type = "orders"

        key_src = "|".join([
            market, sale_date, asin, tracking_id, product_name, category,
            str(qty), f"{revenue:.4f}", f"{commission:.4f}", str(clicks),
        ])
        dedupe_key = hashlib.sha256(key_src.encode("utf-8")).hexdigest()

        out.append({
            "market": market,
            "report_type": report_type,
            "category": category or None,
            "product_name": product_name or None,
            "asin": asin or None,
            "tracking_id": tracking_id or None,
            "sale_date": sale_date or None,
            "qty": qty,
            "returns": returns,
            "revenue": revenue,
            "commission": commission,
            "clicks": clicks,
            "currency": currency,
            "source_file": (filename or "")[:300],
            "dedupe_key": dedupe_key,
        })
    return out


def import_report(
    db: Session,
    filename: str,
    data: bytes,
    default_market: str = "BR",
) -> dict:
    """Le o arquivo e grava as linhas novas (sem duplicar)."""
    parsed = parse_report(filename, data, default_market=default_market)
    imported = 0
    skipped = 0
    markets: dict[str, int] = {}
    for row in parsed:
        exists = (
            db.query(AmazonSale.id)
            .filter(AmazonSale.dedupe_key == row["dedupe_key"])
            .first()
        )
        if exists:
            skipped += 1
            continue
        db.add(AmazonSale(**row))
        imported += 1
        markets[row["market"]] = markets.get(row["market"], 0) + 1
    db.commit()
    return {
        "imported": imported,
        "skipped": skipped,
        "total_rows": len(parsed),
        "markets": markets,
    }


# ----------------------------------------------------------------
# Importacao AUTOMATICA: o ATLAS procura o relatorio sozinho
# ----------------------------------------------------------------
# A Amazon nao deixa puxar as vendas por API. Mas quando o usuario baixa
# o relatorio no site, o arquivo cai na pasta Downloads. Aqui o ATLAS
# vasculha as pastas monitoradas e importa sozinho os relatorios novos,
# para que a pessoa so precise abrir a pagina e ver os numeros.

_MAX_SCAN_BYTES = 15 * 1024 * 1024  # 15 MB
_SCAN_MIN_INTERVAL = 45.0  # segundos entre buscas automaticas
_RECENT_DAYS = 60  # so olha arquivos recentes

# Dicas para reconhecer que um arquivo e mesmo um relatorio da Amazon,
# evitando importar planilhas que nao tem nada a ver.
_AMAZON_FILE_HINTS = (
    "amazon", "associate", "associados", "afiliado", "affiliate",
    "ganhos", "pedidos", "cliques", "earnings", "orders", "fee",
    "tracking", "achadosatlasb", "atlasfindsus", "relatorio", "report",
)
_AMAZON_CONTENT_HINTS = (
    "achadosatlasb", "atlasfindsus", "tracking id", "asin", "ad fees",
    "ganhos", "comiss", "product sales", "items shipped", "product name",
)

_last_scan_ts = 0.0
_seen_files: dict[str, tuple[float, int]] = {}


def _scan_dirs() -> list[Path]:
    """Pastas onde o ATLAS procura os relatorios."""
    dirs: list[Path] = []
    cfg = (getattr(settings, "ATLAS_AMAZON_REPORTS_DIR", "") or "").strip()
    if cfg:
        dirs.append(Path(cfg))
    root = os.environ.get("ATLAS_ROOT") or os.getcwd()
    dirs.append(Path(root) / "relatorios_amazon")
    try:
        dirs.append(Path.home() / "Downloads")
    except Exception:
        pass
    # remove repetidas mantendo a ordem
    out: list[Path] = []
    seen: set[str] = set()
    for d in dirs:
        try:
            key = str(d.resolve())
        except Exception:
            continue
        if key not in seen:
            seen.add(key)
            out.append(d)
    return out


def _looks_like_amazon_file(filename: str, data: bytes) -> bool:
    low = filename.lower()
    if any(h in low for h in _AMAZON_FILE_HINTS):
        return True
    head = data[:8192]
    text = ""
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = head.decode(enc).lower()
            break
        except Exception:
            continue
    return any(h in text for h in _AMAZON_CONTENT_HINTS)


def auto_scan_and_import(db: Session, force: bool = False) -> dict:
    """Procura relatorios da Amazon nas pastas monitoradas e importa
    automaticamente os novos. Seguro chamar sempre: nao duplica dados
    (dedupe_key) e ignora arquivos que nao sejam relatorios da Amazon."""
    global _last_scan_ts
    now = time.time()
    if not force and (now - _last_scan_ts) < _SCAN_MIN_INTERVAL:
        return {"scanned": 0, "imported_files": 0, "imported_rows": 0, "skipped": True}
    _last_scan_ts = now

    # garante a pasta padrao do projeto
    try:
        root = os.environ.get("ATLAS_ROOT") or os.getcwd()
        (Path(root) / "relatorios_amazon").mkdir(exist_ok=True)
    except Exception:
        pass

    cutoff = now - _RECENT_DAYS * 86400
    scanned = 0
    imported_files = 0
    imported_rows = 0

    for d in _scan_dirs():
        try:
            if not d.exists():
                continue
            entries = list(d.iterdir())
        except Exception:
            continue
        for f in entries:
            try:
                if not f.is_file():
                    continue
                low = f.name.lower()
                if not (low.endswith(".csv") or low.endswith(".xlsx") or low.endswith(".txt")):
                    continue
                st = f.stat()
                if st.st_mtime < cutoff:
                    continue
                if st.st_size <= 0 or st.st_size > _MAX_SCAN_BYTES:
                    continue
                key = str(f.resolve())
                sig = (st.st_mtime, st.st_size)
                if not force and _seen_files.get(key) == sig:
                    continue  # ja lido, sem mudancas
                scanned += 1
                data = f.read_bytes()
                if not _looks_like_amazon_file(f.name, data):
                    _seen_files[key] = sig
                    continue
                try:
                    res = import_report(db, f.name, data, default_market="BR")
                except Exception:
                    _seen_files[key] = sig
                    continue
                _seen_files[key] = sig
                if res.get("imported"):
                    imported_files += 1
                    imported_rows += int(res["imported"])
            except Exception:
                continue

    return {
        "scanned": scanned,
        "imported_files": imported_files,
        "imported_rows": imported_rows,
    }



# ----------------------------------------------------------------
# Estatisticas para a pagina "Vendas Amazon"
# ----------------------------------------------------------------
def _empty_totals() -> dict:
    return {"commission": 0.0, "revenue": 0.0, "qty": 0, "clicks": 0, "returns": 0}


def compute_stats(
    db: Session,
    market: Optional[str] = None,
    days: Optional[int] = None,
    top: int = 10,
) -> dict:
    q = db.query(AmazonSale)
    if market in ("BR", "US"):
        q = q.filter(AmazonSale.market == market)
    if days and days > 0:
        cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
        q = q.filter(
            (AmazonSale.sale_date.is_(None)) | (AmazonSale.sale_date >= cutoff)
        )
    rows = q.all()

    totals = _empty_totals()
    by_market: dict[str, dict] = {}
    products: dict[str, dict] = {}
    by_period: dict[str, dict] = {}

    for r in rows:
        totals["commission"] += r.commission or 0.0
        totals["revenue"] += r.revenue or 0.0
        totals["qty"] += r.qty or 0
        totals["clicks"] += r.clicks or 0
        totals["returns"] += r.returns or 0

        bm = by_market.setdefault(r.market, _empty_totals())
        bm["commission"] += r.commission or 0.0
        bm["revenue"] += r.revenue or 0.0
        bm["qty"] += r.qty or 0
        bm["clicks"] += r.clicks or 0
        bm["returns"] += r.returns or 0

        key = (r.asin or r.product_name or "?") + "|" + r.market
        p = products.setdefault(key, {
            "product_name": r.product_name or r.asin or "(sem nome)",
            "asin": r.asin,
            "market": r.market,
            "qty": 0, "commission": 0.0, "revenue": 0.0, "clicks": 0,
        })
        p["qty"] += r.qty or 0
        p["commission"] += r.commission or 0.0
        p["revenue"] += r.revenue or 0.0
        p["clicks"] += r.clicks or 0

        if r.sale_date:
            d = by_period.setdefault(r.sale_date, {
                "date": r.sale_date, "revenue": 0.0, "commission": 0.0, "qty": 0,
            })
            d["revenue"] += r.revenue or 0.0
            d["commission"] += r.commission or 0.0
            d["qty"] += r.qty or 0

    plist = list(products.values())
    top_sold = sorted(plist, key=lambda x: x["qty"], reverse=True)
    top_sold = [p for p in top_sold if p["qty"] > 0][:top]
    top_earnings = sorted(plist, key=lambda x: x["commission"], reverse=True)
    top_earnings = [p for p in top_earnings if p["commission"] > 0][:top]
    top_clicked = sorted(plist, key=lambda x: x["clicks"], reverse=True)
    top_clicked = [p for p in top_clicked if p["clicks"] > 0][:top]

    period = sorted(by_period.values(), key=lambda x: x["date"])

    clicks = totals["clicks"]
    conversion = (totals["qty"] / clicks * 100.0) if clicks > 0 else 0.0

    # Cliques rastreados pelo proprio ATLAS (encurtador /go/).
    internal_clicks = 0
    try:
        from app.models.dashboard import ShortLink
        internal_clicks = int(
            sum((s.clicks or 0) for s in db.query(ShortLink.clicks).all())
        )
    except Exception:  # noqa: BLE001
        internal_clicks = 0

    last = (
        db.query(AmazonSale)
        .order_by(AmazonSale.imported_at.desc())
        .first()
    )
    last_import = None
    if last:
        last_import = {
            "imported_at": last.imported_at.isoformat() if last.imported_at else None,
            "source_file": last.source_file,
        }

    return {
        "has_data": len(rows) > 0,
        "totals": {**totals, "conversion": round(conversion, 2)},
        "by_market": by_market,
        "currency_by_market": _CURRENCY_BY_MARKET,
        "top_sold": top_sold,
        "top_earnings": top_earnings,
        "top_clicked": top_clicked,
        "by_period": period,
        "internal_clicks": internal_clicks,
        "row_count": len(rows),
        "last_import": last_import,
    }
